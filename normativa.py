"""Catálogo de controles transversales y parseo de matrices/checklists."""
import csv
import io
import unicodedata

# Catálogo de controles preventivos transversales (aplican a todos los contratos).
CONTROLES = [
    {'key': 'fatiga_somnolencia', 'label': 'Fatiga y Somnolencia',
     'keywords': ['fatiga', 'somnolencia', 'descanso', 'sueno']},
    {'key': 'vehiculos_livianos', 'label': 'Vehículos Livianos / Conducción',
     'keywords': ['vehiculo liviano', 'vehiculos livianos', 'conduccion', 'preuso', 'camioneta']},
    {'key': 'izaje', 'label': 'Izaje y Levante',
     'keywords': ['izaje', 'grua', 'levante', 'maniobra de izaje', 'aparejo']},
    {'key': 'psicosocial_karin', 'label': 'Psicosocial / Ley Karin',
     'keywords': ['psicosocial', 'ley karin', 'acoso', 'istas', 'riesgo psicosocial']},
    {'key': 'controles_criticos', 'label': 'Controles Críticos (ECF)',
     'keywords': ['control critico', 'controles criticos', 'ecf', 'fatalidad']},
]

CONTROL_LABEL = {c['key']: c['label'] for c in CONTROLES}

# Valores que indican conformidad en la columna de estado.
_CONFORME = {'cumple', 'conforme', 'aprobado', 'aprobada', 'si', 'cerrada', 'cerrado',
             'acreditado', 'ok', 'validado', 'aplica'}


def _norm(s):
    """minúsculas sin acentos."""
    s = str(s or '').strip().lower()
    return ''.join(c for c in unicodedata.normalize('NFD', s)
                   if unicodedata.category(c) != 'Mn')


def _detectar_control(texto):
    t = _norm(texto)
    for c in CONTROLES:
        if any(kw in t for kw in c['keywords']):
            return c['key']
    return None


def _es_conforme(valor):
    return _norm(valor) in _CONFORME


def _idx_columnas(headers):
    """Ubica la columna de requisito/control y la de estado/conformidad."""
    h = [_norm(x) for x in headers]
    col_req = col_est = None
    for i, x in enumerate(h):
        if col_req is None and any(k in x for k in ['requisito', 'control', 'item', 'descripcion', 'medida']):
            col_req = i
        if col_est is None and any(k in x for k in ['estado', 'conformidad', 'cumple', 'brecha', 'cumplimiento']):
            col_est = i
    return col_req, col_est


def _procesar_filas(filas):
    """filas: lista de listas (primera fila = encabezados)."""
    if not filas:
        return []
    headers = filas[0]
    col_req, col_est = _idx_columnas(headers)
    if col_req is None:
        col_req = 0
    detectados = {}
    for fila in filas[1:]:
        if not fila:
            continue
        texto = fila[col_req] if col_req < len(fila) else ''
        key = _detectar_control(texto)
        if not key:
            continue
        conforme = True
        if col_est is not None and col_est < len(fila):
            conforme = _es_conforme(fila[col_est])
        # un control se marca conforme si alguna fila asociada lo está
        detectados[key] = detectados.get(key, False) or conforme
    return [{'control_key': k, 'conforme': v, 'texto': CONTROL_LABEL[k]}
            for k, v in detectados.items()]


def parsear_matriz(filestorage):
    """Recibe un werkzeug FileStorage; devuelve [{control_key, conforme, texto}]."""
    nombre = (getattr(filestorage, 'filename', '') or '').lower()
    data = filestorage.read()
    try:
        if nombre.endswith('.csv'):
            texto = data.decode('utf-8-sig', errors='replace')
            # detecta separador ; o ,
            sep = ';' if texto.count(';') >= texto.count(',') else ','
            filas = list(csv.reader(io.StringIO(texto), delimiter=sep))
            return _procesar_filas(filas)
        if nombre.endswith('.xlsx'):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            filas = [[('' if v is None else v) for v in row]
                     for row in ws.iter_rows(values_only=True)]
            return _procesar_filas(filas)
    except Exception:
        return []
    return []
