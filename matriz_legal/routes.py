"""Módulo de Legalidad — Matriz Legal D.S. 44. Rutas aisladas en Blueprint.

Sin url_prefix a propósito: las URLs /api/matriz-legal* se conservan idénticas porque el JS del
dashboard (cargarTrazabilidad) ya depende de ellas.
"""
from io import BytesIO

from flask import Blueprint, render_template, request, jsonify, session

import db
import cumplimiento
from core_auth import login_required, empresa_required, onboarding_required, empresa_id

from . import service
from formato import normalizar_control_operativo

bp = Blueprint('matriz_legal', __name__, template_folder='templates')


@bp.route('/matriz-legal')
@empresa_required
@onboarding_required
def panel():
    """Cerebro Legal: Matriz Legal editable. La trazabilidad del FUF 44 (solo lectura) vive ahora
    en Cumplimiento DS 44, junto con el modal donde se responde — no se duplica aquí."""
    eid = empresa_id()
    service.asegurar_core(eid)
    return render_template('matriz_legal/panel.html',
                           empresa=db.empresa_de(session.get('rut'), eid),
                           pilares=cumplimiento.PILARES,
                           pilares_cortos=cumplimiento.PILARES_CORTOS)


@bp.route('/api/matriz-legal', methods=['GET'])
@empresa_required
def api_matriz_get():
    eid = empresa_id()
    service.asegurar_core(eid)
    return jsonify(db.matriz_legal(eid))


@bp.route('/api/matriz-legal/resumen', methods=['GET'])
@empresa_required
@onboarding_required
def api_matriz_resumen():
    """% de cumplimiento de la Matriz Legal (para el gráfico de la portada del panel, OBS-3A)."""
    eid = empresa_id()
    service.asegurar_core(eid)
    return jsonify(db.matriz_legal_resumen(eid))


@bp.route('/api/matriz-legal/analizar-ley', methods=['POST'])
@empresa_required
@onboarding_required
def api_matriz_analizar_ley():
    """Analiza una ley (link BCN o texto) con IA y sugiere los campos del requisito (OBS-3B)."""
    import ia
    f = request.get_json(silent=True) or {}
    return jsonify(ia.analizar_ley(f.get('texto') or f.get('link') or ''))


@bp.route('/api/matriz-legal/catalogo', methods=['GET'])
@empresa_required
@onboarding_required
def api_matriz_catalogo():
    """Catálogo transversal SST agrupado por pilar, marcando lo ya insertado (para el modal)."""
    return jsonify(db.catalogo_legal_estado(empresa_id()))


@bp.route('/api/matriz-legal/insertar', methods=['POST'])
@empresa_required
@onboarding_required
def api_matriz_insertar():
    """Inserta uno o varios requisitos del catálogo transversal en la Matriz Legal."""
    f = request.get_json(silent=True) or {}
    codigos = f.get('codigos') or ([f['codigo']] if f.get('codigo') else [])
    eid, insertados, ya = empresa_id(), 0, 0
    for cod in codigos:
        r = db.insertar_requisito_legal(eid, cod)
        if r is None:
            return jsonify({'error': f'Requisito «{cod}» no está en el catálogo.'}), 404
        if r.get('ya_existe'):
            ya += 1
        else:
            insertados += 1
    return jsonify({'ok': True, 'insertados': insertados, 'ya_estaban': ya,
                    'matriz': db.matriz_legal(eid)})


@bp.route('/api/matriz-legal', methods=['POST'])
@empresa_required
@onboarding_required
def api_matriz_guardar():
    data = request.get_json(silent=True) or {}
    # Capa Operativa: si es fila nueva sin ID, genera uno (OP-N) automáticamente.
    if not (data.get('id_requisito') or '').strip():
        existentes = db.matriz_legal(empresa_id())
        n = 1 + sum(1 for r in existentes if str(r.get('id_requisito') or '').startswith('OP-'))
        data['id_requisito'] = f'OP-{n:03d}'
        data.setdefault('capa', 'operativa')
    if 'control_operativo' in data:
        data['control_operativo'] = normalizar_control_operativo(data['control_operativo'])
    return jsonify(db.requisito_guardar(empresa_id(), data))


@bp.route('/api/matriz-legal/<int:rid>/eliminar', methods=['POST'])
@empresa_required
@onboarding_required
def api_matriz_eliminar(rid):
    res = db.requisito_eliminar(empresa_id(), rid)
    if res.get('error'):
        return jsonify(res), 400
    return jsonify(db.matriz_legal(empresa_id()))


@bp.route('/api/matriz-legal/importar', methods=['POST'])
@empresa_required
@onboarding_required
def api_matriz_importar():
    """Importa la Matriz Legal desde un Excel normalizado (columnas ID_Requisito, Origen,
    Cuerpo_Normativo, Requisito_Legal, Riesgo_Asociado, Control_Operativo, Frecuencia,
    Estado_Avance). Reusa openpyxl."""
    archivo = request.files.get('archivo')
    if not archivo or not archivo.filename:
        return jsonify({'error': 'No se recibió archivo.'}), 400
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(archivo.read()), data_only=True)
        ws = wb.active
        filas = list(ws.iter_rows(values_only=True))
    except Exception as e:      # noqa: BLE001
        return jsonify({'error': f'No se pudo leer el Excel: {type(e).__name__}.'}), 400
    if not filas:
        return jsonify({'error': 'El archivo está vacío.'}), 400
    encabezados = [str(h or '').strip().lower().replace(' ', '_') for h in filas[0]]
    campo = {'id_requisito': 'id_requisito', 'origen': 'origen', 'cuerpo_normativo': 'cuerpo_normativo',
             'requisito_legal': 'requisito_legal', 'riesgo_asociado': 'riesgo_asociado',
             'control_operativo': 'control_operativo', 'frecuencia': 'frecuencia',
             'estado_avance': 'estado_avance', 'responsable': 'responsable', 'capa': 'capa',
             'categoria': 'categoria'}
    n = 0
    for fila in filas[1:]:
        data = {}
        for i, h in enumerate(encabezados):
            if h in campo and i < len(fila) and fila[i] is not None:
                data[campo[h]] = str(fila[i]).strip()
        if not data:
            continue
        if data.get('control_operativo'):
            data['control_operativo'] = normalizar_control_operativo(data['control_operativo'])
        db.requisito_guardar(empresa_id(), data)
        n += 1
    return jsonify({'ok': True, 'importados': n, 'matriz': db.matriz_legal(empresa_id())})


@bp.route('/api/matriz-legal/<int:rid>/riesgos', methods=['GET'])
@empresa_required
def api_matriz_riesgos(rid):
    """Legal → MIPER: actividades/riesgos vinculados a un requisito legal."""
    return jsonify(db.riesgos_de_requisito(empresa_id(), rid))
