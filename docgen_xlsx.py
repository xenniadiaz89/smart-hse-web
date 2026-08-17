"""Exportación de la Matriz de Riesgos (MIPER) al formato Excel estándar del mandante.

Toma la plantilla vendorizada (con su cabecera, encabezados de columna y celdas combinadas) y la
rellena con los datos del contrato (cabecera) y los RiesgoItem de la faena (filas de la matriz),
para que el asesor descargue el Excel EN EL MISMO ORDEN/LAYOUT del mandante y lo suba a su nube.
Reusa openpyxl (ya en requirements). Salida: bytes .xlsx.

Datos de cabecera = los mismos que se piden al ingresar el contrato (contrato.datos_json). Los campos que
la app aún no modela (evaluación residual, puesto, responsable) quedan en blanco para que el asesor los
complete.
"""
import json
import os
from io import BytesIO

import iper

import openpyxl

_TPL_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'plantillas')
_TPL_MIPER = os.path.join(_TPL_DIR, 'SIGO-F-006_MIPER.xlsx')
_TPL_MAPA = os.path.join(_TPL_DIR, 'SIGO-F-011_MapaProceso.xlsx')

# Hoja base (proceso) que se usa como layout de la matriz; las demás hojas de proceso se descartan.
_HOJA_DATOS = 'Conducción'
_HOJAS_PROCESO_EXTRA = ('Anexo N° 1', 'Trabajo administrativo')
_FILA_DATOS = 10          # primera fila de datos de la matriz (bajo el doble encabezado 8-9)
_MAX_COL = 20             # columna T (RESPONSABLE)


def _datos(contrato):
    """datos_json del contrato como dict (tolerante)."""
    try:
        return json.loads(contrato.get('datos_json') or '{}')
    except (TypeError, ValueError):
        return {}


def build_miper_xlsx(contrato, riesgos):
    """Devuelve los bytes de un .xlsx SIGO-F-006 con la cabecera del contrato y una fila por riesgo."""
    wb = openpyxl.load_workbook(_TPL_MIPER)
    ws = wb[_HOJA_DATOS] if _HOJA_DATOS in wb.sheetnames else wb[wb.sheetnames[0]]
    ws.title = 'MIPER'
    # Descartar las otras hojas de proceso con datos de ejemplo (conserva hojas de catálogo).
    for nombre in _HOJAS_PROCESO_EXTRA:
        if nombre in wb.sheetnames:
            del wb[nombre]

    d = _datos(contrato)
    division = contrato.get('mandante') or 'Mandante'
    empresa = contrato.get('empresa') or ''
    numero = contrato.get('numero') or ''
    # Cabecera fila 6 (celdas combinadas B6:F7 / G6:L7 / M6:P7): se rellena junto al rótulo.
    ws['B6'] = f"MANDANTE: {division}"
    ws['G6'] = f"GERENCIA: {d.get('gerencia', '')}"
    ws['M6'] = (f"SUPERINTENDENCIA / EE.CC / N° CONTRATO: "
                f"{d.get('superintendencia', '')} / {empresa} / {numero}")

    # Limpiar el área de datos (desde fila 10): quitar combinaciones y valores del ejemplo.
    for rango in list(ws.merged_cells.ranges):
        if rango.min_row >= _FILA_DATOS:
            ws.unmerge_cells(str(rango))
    for fila in ws.iter_rows(min_row=_FILA_DATOS, max_row=ws.max_row, min_col=1, max_col=_MAX_COL):
        for celda in fila:
            celda.value = None

    # Escribir una fila por riesgo. Campos ausentes en el modelo → en blanco (los completa el asesor).
    r = _FILA_DATOS
    for x in riesgos:
        tarea = x.get('tarea') or ''
        proceso = x.get('proceso') or (tarea.split('·')[0].strip() if '·' in tarea else (tarea or 'Faena'))
        codigo = ' · '.join(p for p in [x.get('ecf_punto'), x.get('riesgo')] if p)
        ws.cell(r, 2, proceso)                       # B PROCESO
        ws.cell(r, 3, tarea)                         # C ACTIVIDAD / TRABAJO
        ws.cell(r, 4, tarea)                         # D TAREA
        ws.cell(r, 5, x.get('puesto'))               # E PUESTO (Ronda 25: TareaIPER.puesto)
        # F N° PERSONAS → en blanco (dotación por puesto no modelada)
        ws.cell(r, 7, x.get('peligro'))              # G IDENTIFICACIÓN DE PELIGROS
        ws.cell(r, 8, iper.GEMA.get(x.get('gema'), ''))   # H CATEGORÍA (Ronda 25: GEMA)
        ws.cell(r, 9, codigo)                        # I CÓDIGO - RIESGO ESPECÍFICO
        ws.cell(r, 10, x.get('probabilidad'))        # J PROBABILIDAD (inherente)
        ws.cell(r, 11, x.get('consecuencia'))        # K CONSECUENCIA
        ws.cell(r, 12, x.get('vep'))                 # L VEP / MAGNITUD
        ws.cell(r, 13, x.get('nivel_riesgo'))        # M NIVEL DE RIESGO
        ws.cell(r, 14, x.get('medida_control'))      # N MEDIDAS PREVENTIVAS
        ws.cell(r, 15, iper.TIPOS_CONTROL.get(x.get('tipo_control'), x.get('tipo_control')))  # O PRELACIÓN
        # P..S EVALUACIÓN RESIDUAL (Ronda 25) — el residual que deja el control validado
        ws.cell(r, 16, x.get('probabilidad_residual'))    # P PROBABILIDAD residual
        ws.cell(r, 17, x.get('consecuencia_residual'))    # Q CONSECUENCIA residual
        ws.cell(r, 18, x.get('vep_residual'))             # R VEP residual
        ws.cell(r, 19, x.get('nivel_riesgo_residual'))    # S NIVEL DE RIESGO residual
        # T RESPONSABLE → en blanco (lo completa el asesor)
        r += 1

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ── Mapa de Proceso SIGO-F-011 ──────────────────────────────────────────────
_MAPA_HOJA = 'EJEMPLO MP'
_MAPA_FILA_ACT = 14       # primera fila de la tabla Actividades/Tareas (bajo el encabezado 12-13)


def build_mapa_xlsx(contrato, tareas):
    """Devuelve los bytes de un .xlsx SIGO-F-011 (Mapa de Proceso) con Antecedentes autocompletados,
    la lista de Procesos y la tabla Actividades→Tareas del contrato. Campos no modelados (cargo, lugar,
    dotación) quedan en blanco para que el asesor los complete."""
    wb = openpyxl.load_workbook(_TPL_MAPA)
    ws = wb[_MAPA_HOJA] if _MAPA_HOJA in wb.sheetnames else wb[wb.sheetnames[0]]
    ws.title = 'Mapa de Proceso'

    d = _datos(contrato)
    division = contrato.get('mandante') or 'Mandante'
    empresa = contrato.get('empresa') or ''
    numero = contrato.get('numero') or ''
    # A. Antecedentes (valores en fila 5, bajo las etiquetas de fila 4).
    ws['B5'] = division                                   # Centro de Trabajo
    ws['E5'] = d.get('gerencia', '')                      # Gerencia
    ws['G5'] = d.get('superintendencia', '')              # Superintendencia / Dirección
    ws['I5'] = d.get('area', '')                          # Área
    ws['L5'] = f"{empresa} / {numero}"                    # EE.CC / N° CONTRATO
    # B. Reseña
    servicio = d.get('nombre_servicio', '')
    ws['B7'] = (f"Reseña: {empresa} presta el servicio "
                f"{servicio or '(servicio)'} para {division}.")

    procesos = []
    for t in tareas:
        p = t.get('proceso') or 'Proceso'
        if p not in procesos:
            procesos.append(p)
    # C. Procesos operativos: se listan numerados en la celda de la sección (fila 10).
    ws['B10'] = len(procesos) or 1
    ws['C10'] = '\n'.join(f"{i}. {p}" for i, p in enumerate(procesos, 1)) or '(sin procesos)'

    # D. Actividades → Tareas: limpiar el área de datos (quitar combinaciones y ejemplo) y escribir.
    for rango in list(ws.merged_cells.ranges):
        if rango.min_row >= _MAPA_FILA_ACT:
            ws.unmerge_cells(str(rango))
    for fila in ws.iter_rows(min_row=_MAPA_FILA_ACT, max_row=max(ws.max_row, _MAPA_FILA_ACT), min_col=1, max_col=14):
        for celda in fila:
            celda.value = None
    r = _MAPA_FILA_ACT
    for i, t in enumerate(tareas, 1):
        ws.cell(r, 2, f"{procesos.index(t.get('proceso') or 'Proceso') + 1}.{i}")  # B N° Actividad
        ws.cell(r, 3, t.get('proceso'))                  # C Actividad del Proceso
        ws.cell(r, 5, f"{i}")                            # E N° Tarea
        ws.cell(r, 6, t.get('tarea'))                    # F Tarea de la Actividad
        # H R/NR, I Cargo, J Lugar, K/L/M/N dotación → en blanco (los completa el asesor)
        r += 1

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
