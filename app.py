import os
import json
import re
from io import BytesIO
from datetime import date
from functools import wraps

from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, jsonify, send_file)
from werkzeug.security import generate_password_hash, check_password_hash

from models import sqla
import db
import normativa
import resso
import ia
import correccion
import cumplimiento
import alertas
import docgen
import iper

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'smarthse-dev-key-cambiar-en-render')

# ── Base de datos: PostgreSQL en producción (DATABASE_URL), SQLite en local ──
_db_url = os.environ.get('DATABASE_URL', '')
if _db_url.startswith('postgres://'):          # Render entrega 'postgres://'; SQLAlchemy pide 'postgresql://'
    _db_url = _db_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = _db_url or 'sqlite:///' + os.path.join(
    os.path.dirname(__file__), 'smarthse.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
sqla.init_app(app)

with app.app_context():
    db.init_db()                                # crea tablas (auto-migración) + siembra mapping


# ─────────────────────────── Utilidades RUT / clave ────────────────────────
def normalizar_rut(rut):
    """Quita puntos y guion, deja el dígito verificador en mayúscula."""
    r = re.sub(r'[^0-9kK]', '', rut or '').upper()
    if len(r) < 2:
        return r
    return r[:-1] + '-' + r[-1]


def rut_valido(rut):
    """Valida el dígito verificador chileno (módulo 11)."""
    r = re.sub(r'[^0-9kK]', '', rut or '').upper()
    if len(r) < 2:
        return False
    cuerpo, dv = r[:-1], r[-1]
    if not cuerpo.isdigit():
        return False
    suma, factor = 0, 2
    for d in reversed(cuerpo):
        suma += int(d) * factor
        factor = 2 if factor == 7 else factor + 1
    resto = 11 - (suma % 11)
    dv_calc = 'K' if resto == 10 else '0' if resto == 11 else str(resto)
    return dv == dv_calc


def clave_valida(c):
    """Alfanumérica: mín. 6 caracteres con al menos una letra y un dígito."""
    return bool(c) and len(c) >= 6 and re.search(r'[A-Za-z]', c) and re.search(r'\d', c)


def normalizar_id(valor):
    """Normaliza el N° SNS para usarlo como clave de cuenta (sin puntos/guion, mayúsculas)."""
    return re.sub(r'[^0-9A-Za-z]', '', valor or '').upper()


# ─────────────────────────── Control de acceso ─────────────────────────────
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get('rut'):
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Sesión expirada. Reingresa a la consola.'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrapper


def _empresa_id():
    """Empresa en foco (Ronda 12). None si el asesor aún no seleccionó una."""
    return session.get('empresa_id')


def empresa_required(f):
    """Exige empresa activa. Para /api → 409 JSON; para vistas → redirige a Mis Empresas."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get('rut'):
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Sesión expirada. Reingresa a la consola.'}), 401
            return redirect(url_for('login'))
        if not session.get('empresa_id'):
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Selecciona o registra una empresa primero.'}), 409
            return redirect(url_for('empresas'))
        return f(*args, **kwargs)
    return wrapper


# ─────────────────────────────── Rutas ─────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Acceso exclusivo con RUT + clave."""
    if request.method == 'POST':
        rut_raw = request.form.get('rut', '')
        clave = request.form.get('clave', '')
        key = normalizar_rut(rut_raw)
        u = db.usuario_get(key)
        if not u or not (u.get('pass_hash') and check_password_hash(u['pass_hash'], clave)):
            return render_template('login.html', error='RUT o clave incorrectos.', rut=rut_raw)
        session['rut'] = key                 # llave de cuenta = RUT normalizado
        session['sns'] = u.get('sns') or ''  # ID profesional visible
        session['nombre'] = u.get('nombre')
        session['rol'] = u.get('rol', 'asesor')
        return redirect(url_for('dashboard'))
    return render_template('login.html')


@app.route('/registro', methods=['GET', 'POST'])
def registro():
    """Registro con RUT + clave + SNS (ID profesional) + nombre. El SNS se pide solo aquí."""
    if request.method == 'POST':
        f = request.form
        nombre = (f.get('nombre', '')).strip()
        rut_raw = (f.get('rut', '')).strip()
        sns = (f.get('sns', '')).strip()
        clave = f.get('clave', '')
        datos = {'nombre': nombre, 'rut': rut_raw, 'sns': sns}

        if not (nombre and rut_raw and sns):
            return render_template('registro.html', error='Completa nombre, RUT y N° SNS.', **datos)
        if not rut_valido(rut_raw):
            return render_template('registro.html', error='El RUT ingresado no es válido.', **datos)
        if not clave_valida(clave):
            return render_template('registro.html',
                                   error='La clave debe ser alfanumérica de al menos 6 caracteres (con letras y números).', **datos)

        key = normalizar_rut(rut_raw)
        if db.usuario_get(key):
            return render_template('registro.html', error='Ya existe una cuenta con ese RUT.', **datos)

        db.usuario_crear(key, rut_raw, sns, nombre, rol='asesor',
                         pass_hash=generate_password_hash(clave))
        session['rut'] = key
        session['sns'] = sns
        session['nombre'] = nombre
        session['rol'] = 'asesor'
        session.pop('empresa_id', None)
        # Ronda 18: entra directo a la consola; la empresa se registra dentro de "Mis Contratos".
        return redirect(url_for('dashboard'))
    return render_template('registro.html')


@app.route('/prueba')
def prueba():
    """Acceso de PRUEBA momentáneo (sin pago). Identidad demo FIJA ('DEMO') y una
    empresa demo estable para que el workspace sobreviva reinicios/cookies perdidas.
    TEMPORAL: cuando se active el cobro, se reemplaza por el flujo de pago."""
    session['rut'] = 'DEMO'
    session['sns'] = 'DEMO'
    session['nombre'] = 'Usuario de Prueba'
    session['rol'] = 'asesor'
    emps = db.empresas_de('DEMO')
    if emps:
        session['empresa_id'] = emps[0]['id']
    else:
        # Marca blanca (Ronda 21): empresa demo neutra, sin datos corporativos de terceros.
        session['empresa_id'] = db.crear_empresa('DEMO', 'Empresa Demo (Smart HSE)', rubro='Servicios')
    return redirect(url_for('dashboard'))


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))


# ── Empresas (Ronda 18: gestión dentro de la consola, en "Mis Contratos") ──
@app.route('/empresas')
@login_required
def empresas():
    """Compatibilidad: la pantalla /empresas se retiró; ahora la empresa se gestiona en la
    consola ("Mis Contratos"). Redirige al dashboard."""
    return redirect(url_for('dashboard'))


@app.route('/empresas/<int:eid>/seleccionar')
@login_required
def empresa_seleccionar(eid):
    if db.empresa_de(session['rut'], eid):
        session['empresa_id'] = eid
    return redirect(url_for('dashboard'))


@app.route('/dashboard')
@login_required
def dashboard():
    """Consola de Gestión Operativa. Si aún no hay empresa activa, entra igual: 'Mis Contratos'
    muestra el alta de empresa (Ronda 18). Si hay empresas pero ninguna activa, toma la primera."""
    emp = None
    eid = session.get('empresa_id')
    if eid:
        emp = db.empresa_de(session['rut'], eid)
        if not emp:
            session.pop('empresa_id', None)
    if not emp:
        emps = db.empresas_de(session['rut'])
        if emps:
            session['empresa_id'] = emps[0]['id']
            emp = emps[0]
    return render_template('dashboard.html', nombre=session.get('nombre'),
                           sns=session.get('sns'), rol=session.get('rol'), empresa=emp)


# ── API de empresas (gestión desde la consola) ──
@app.route('/api/empresas', methods=['GET'])
@login_required
def api_empresas_get():
    return jsonify({'empresas': db.empresas_de(session['rut']),
                    'activa': session.get('empresa_id')})


# ── Límites del Plan Básico del Asesor (Ronda 20) ──
MAX_EMPRESAS_BASICO = 30
MAX_TRABAJADORES_BASICO = 20      # sobre 20 → plan de 25-50; más arriba, otro pack


@app.route('/api/plan', methods=['GET'])
@login_required
def api_plan():
    """Uso del Plan Básico: empresas de la cuenta y trabajadores de la empresa activa."""
    empresas = db.empresas_de(session['rut'])
    eid = session.get('empresa_id')
    trabajadores = len(db.trabajadores_de(eid)) if eid else 0
    return jsonify({'plan': 'Básico', 'empresas': len(empresas), 'max_empresas': MAX_EMPRESAS_BASICO,
                    'trabajadores': trabajadores, 'max_trabajadores': MAX_TRABAJADORES_BASICO})


@app.route('/api/empresas', methods=['POST'])
@login_required
def api_empresa_crear():
    f = request.get_json(silent=True) or request.form
    razon = (f.get('razon_social') or '').strip()
    if not razon:
        return jsonify({'error': 'Indica la Razón Social.'}), 400
    if len(db.empresas_de(session['rut'])) >= MAX_EMPRESAS_BASICO:
        return jsonify({'error': 'limite_empresas',
                        'mensaje': f'Alcanzaste el límite del Plan Básico ({MAX_EMPRESAS_BASICO} '
                                   'empresas). Migra a un pack corporativo superior para gestionar más.'}), 403
    eid = db.crear_empresa(
        session['rut'], razon,
        rut_empresa=(f.get('rut_empresa') or '').strip() or None,
        mutual=(f.get('mutual') or '').strip() or None,
        n_adherente=(f.get('n_adherente') or '').strip() or None,
        rubro=(f.get('rubro') or '').strip() or None)
    session['empresa_id'] = eid
    return jsonify({'ok': True, 'empresa_id': eid, 'empresas': db.empresas_de(session['rut'])})


@app.route('/api/empresas/<int:eid>/seleccionar', methods=['POST'])
@login_required
def api_empresa_seleccionar(eid):
    if not db.empresa_de(session['rut'], eid):
        return jsonify({'error': 'Empresa no encontrada.'}), 404
    session['empresa_id'] = eid
    return jsonify({'ok': True, 'empresa_id': eid})


# ── Adhesión a Mutualidad (Ley 16.744) — nivel empresa, una vez por RUT ──
@app.route('/api/empresa/adhesion', methods=['GET'])
@empresa_required
def api_adhesion_get():
    return jsonify(db.adhesion_estado(session['rut'], _empresa_id()))


@app.route('/api/empresa/adhesion', methods=['POST'])
@empresa_required
def api_adhesion_set():
    f = request.get_json(silent=True) or request.form
    db.adhesion_guardar(session['rut'], _empresa_id(),
                        mutual=(f.get('mutual') or '').strip() or None,
                        n_adherente=(f.get('n_adherente') or '').strip() or None)
    return jsonify(db.adhesion_estado(session['rut'], _empresa_id()))


@app.route('/api/empresa/adhesion/<tipo>', methods=['POST'])
@empresa_required
def api_adhesion_cert(tipo):
    if tipo not in ('adhesion', 'siniestralidad', 'cotizaciones'):
        return jsonify({'error': 'Tipo de certificado inválido.'}), 400
    archivo = request.files.get('archivo')
    if not archivo or not archivo.filename:
        return jsonify({'error': 'No se recibió archivo.'}), 400
    rut, eid = session['rut'], _empresa_id()
    emp = db.empresa_de(rut, eid)
    base_cid = db.contrato_base(eid, rut, emp.get('razon_social'))
    doc_id = db.registrar_documento(base_cid, f'cert_{tipo}_{archivo.filename}', 'Adhesión',
                                    f'cert_{tipo}', contenido=archivo.read(),
                                    mimetype=archivo.mimetype or 'application/pdf')
    db.adhesion_set_doc(eid, tipo, doc_id)
    return jsonify(db.adhesion_estado(rut, eid))


@app.route('/contratistas')
def contratistas():
    return render_template('contratistas.html')


@app.route('/legislacion')
def legislacion():
    return render_template('legislacion.html')


# ─────────────────── Motor de cumplimiento: contratos / matrices ────────────
def _consolidar(rut, empresa_id=None):
    """Contratos de la empresa activa del asesor, con su estado consolidado.
    Excluye el 'contrato base' (BASE-*), contenedor interno de docs de la Capa Legal."""
    contratos = [c for c in db.listar_contratos(rut, empresa_id if empresa_id is not None else _empresa_id())
                 if not str(c.get('numero', '')).startswith('BASE-')]
    por_num = {c['id']: c['numero'] for c in contratos}
    salida = []
    for c in contratos:
        estados = db.estados_de_contrato(c['id'])
        controles = []
        for ctrl in normativa.CONTROLES:
            e = estados.get(ctrl['key'])
            estado = e['estado'] if e else 'pendiente'
            origen = por_num.get(e['origen_contrato_id']) if e and e.get('origen_contrato_id') else None
            controles.append({'key': ctrl['key'], 'label': ctrl['label'],
                              'estado': estado, 'origen': origen})
        cerradas = sum(1 for x in controles if x['estado'] in ('aprobado', 'acreditado'))
        try:
            datos = json.loads(c['datos_json']) if c.get('datos_json') else {}
        except (TypeError, ValueError):
            datos = {}
        es_codelco = resso.es_codelco(c['mandante'])
        carpeta = None
        if es_codelco:
            car = _carpeta(c['id'])
            cumple = sum(1 for i in car['items'] if i['estado'] == 'cumple')
            na = sum(1 for i in car['items'] if i['estado'] == 'na')
            total = len(car['items'])
            carpeta = {'pct': car['cumplimiento_pct'], 'total': total, 'cumple': cumple,
                       'na': na, 'pendientes': total - cumple - na}
        salida.append({**c, 'controles': controles,
                       'documentos': db.documentos_de(c['id']),
                       'datos': datos, 'es_codelco': es_codelco, 'carpeta': carpeta,
                       'cerradas': cerradas, 'pendientes': len(controles) - cerradas})
    return salida


def _carpeta(cid):
    """Estado de la Carpeta de Arranque (29 ítems) de un contrato."""
    estados = db.estados_carpeta(cid)
    docs = db.docs_por_item(cid)
    items = []
    for it in resso.carpeta_lista():
        e = estados.get(it['n'], {})
        items.append({**it,
                      'estado': e.get('estado', 'pendiente'),
                      'observacion': e.get('observacion', '') or '',
                      'fecha_compromiso': e.get('fecha_compromiso', '') or '',
                      'docs': [{'id': d['id'], 'nombre': d['nombre'], 'tipo': d.get('tipo', 'evidencia')}
                               for d in docs.get(it['n'], [])]})
    aplicables = [i for i in items if i['estado'] != 'na']
    cumple = [i for i in aplicables if i['estado'] == 'cumple']
    pct = round(len(cumple) / len(aplicables) * 100) if aplicables else 0
    return {'items': items, 'cumplimiento_pct': pct}


def _datos(contrato):
    if contrato and contrato.get('datos_json'):
        try:
            return json.loads(contrato['datos_json'])
        except (TypeError, ValueError):
            return {}
    return {}


def _logo_doc(cid):
    """Devuelve el documento-logo (tipo='logo') más reciente del contrato, o None."""
    logos = [d for d in db.documentos_de(cid) if d.get('tipo') == 'logo']
    return logos[0] if logos else None


def _logo_data_uri(rut, cid):
    """Data URI del logo de la empresa (leído desde la BD), o None."""
    doc = _logo_doc(cid)
    if not doc:
        return None
    blob = db.documento_contenido(rut, doc['id'])
    if not blob:
        return None
    import base64
    contenido, mimetype, _ = blob
    b64 = base64.b64encode(contenido).decode()
    return f"data:{mimetype or 'image/png'};base64,{b64}"


def carta_na_html(rut, contrato, datos, item, fundamento):
    """Genera una Carta de No Aplica (N/A) en HTML autocontenido (con logo si existe)."""
    hoy = date.today().strftime('%d-%m-%Y')
    empresa = (datos.get('empresa_contratista') or contrato.get('empresa') or '').strip()
    fund = (fundamento.strip() if fundamento and fundamento.strip() else '[Pendiente de fundamentar]')
    mandante = contrato.get('mandante', '') + (f" — {contrato.get('faena')}" if contrato.get('faena') else '')
    logo = _logo_data_uri(rut, contrato.get('id'))
    logo_html = f'<img src="{logo}" alt="Logo" style="max-height:80px;max-width:220px">' if logo else \
        f'<div style="font-weight:800;color:#006a9b;font-size:20px">{empresa or "Empresa Contratista"}</div>'
    def esc(s):
        return (str(s or '')).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    return f"""<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<title>Carta N/A · Ítem {item['n']} · Contrato {esc(contrato.get('numero',''))}</title>
<style>
 body{{font-family:Arial,Helvetica,sans-serif;color:#1a2b3c;max-width:820px;margin:24px auto;padding:0 24px;line-height:1.5}}
 .head{{display:flex;justify-content:space-between;align-items:center;border-bottom:3px solid #006a9b;padding-bottom:14px;margin-bottom:20px}}
 .titulo{{text-align:right}} h1{{font-size:18px;margin:0;color:#006a9b}} .sub{{font-size:12px;color:#666}}
 table{{width:100%;border-collapse:collapse;margin:14px 0;font-size:13px}}
 td{{padding:6px 8px;border-bottom:1px solid #eee}} td.k{{color:#666;width:230px;font-weight:600}}
 .item{{background:#f4f7f6;border-radius:8px;padding:12px 14px;margin:14px 0}}
 .decl{{margin:18px 0}} .fund{{background:#fffbe6;border:1px solid #ffe58f;border-radius:8px;padding:12px 14px;font-weight:600}}
 .firma{{margin-top:48px;border-top:1px solid #333;width:320px;padding-top:6px;font-size:13px}}
 .pie{{margin-top:24px;font-size:11px;color:#999}}
</style></head><body>
 <div class="head">{logo_html}<div class="titulo"><h1>CARTA DE NO APLICABILIDAD (N/A)</h1><div class="sub">DS 44/2024 · Gestión Preventiva de Riesgos Laborales</div></div></div>
 <table>
  <tr><td class="k">Fecha</td><td>{esc(hoy)}</td></tr>
  <tr><td class="k">Empresa Contratista</td><td>{esc(empresa)}</td></tr>
  <tr><td class="k">N° de Contrato</td><td>{esc(contrato.get('numero',''))}</td></tr>
  <tr><td class="k">Mandante / División</td><td>{esc(mandante)}</td></tr>
  {f'<tr><td class="k">Administrador de Contrato CODELCO</td><td>{esc(datos.get("admin_codelco"))}</td></tr>' if datos.get('admin_codelco') else ''}
 </table>
 <div class="item"><b>Ítem N° {item['n']}: {esc(item['titulo'])}</b><br><span class="sub">Evidencia normalmente requerida: {esc(item['evidencia'])}</span></div>
 <div class="decl"><b>DECLARACIÓN DE NO APLICABILIDAD</b><br>
  Por medio de la presente, la Empresa Contratista declara que el requisito individualizado
  <b>NO APLICA</b> al presente contrato, por el siguiente fundamento:</div>
 <div class="fund">Fundamento: {esc(fund)}</div>
 <p>Esta declaración se incorpora al sistema de gestión preventiva de la empresa (DS 44/2024) para efectos de acreditación y auditoría.</p>
 <div class="firma">{esc(datos.get('experto_eecc') or 'Experto en Prevención de Riesgos EE.CC.')}<br>
  <span class="sub">Experto en Prevención de Riesgos — Empresa Contratista</span></div>
 <div class="pie">Smart HSE Chile · Documento generado automáticamente</div>
</body></html>"""


def generar_carta_na(rut, cid, contrato, n, fundamento):
    """Crea/actualiza la carta N/A (HTML con logo) del ítem n, guardándola en la BD."""
    item = resso.CARPETA_DICT.get(n)
    if not item:
        return
    datos = _datos(contrato)
    html = carta_na_html(rut, contrato, datos, item, fundamento)
    numero = re.sub(r'[^\w-]+', '_', contrato.get('numero', '') or '')
    nombre = f"Carta_NA_item{n:02d}_{numero}.html"
    db.eliminar_doc_tipo(cid, n, 'carta_na')          # evitar duplicados
    db.registrar_documento(cid, nombre, 'N/A', 'carta_na', item_n=n,
                           contenido=html.encode('utf-8'), mimetype='text/html')


def _replicar_lista(rut):
    return db.listar_contratos(rut, _empresa_id())


def replicar_controles(rut, control_key, origen_contrato_id):
    """Hereda como 'acreditado' un control aprobado a los demás contratos de la empresa."""
    for c in _replicar_lista(rut):
        if c['id'] == origen_contrato_id:
            continue
        if db.estado_control(c['id'], control_key) == 'aprobado':
            continue  # no degradar una aprobación propia
        db.set_estado_control(rut, c['id'], control_key, 'acreditado', origen_contrato_id)


@app.route('/api/contratos', methods=['GET'])
@empresa_required
def api_contratos():
    return jsonify(_consolidar(session['rut']))


@app.route('/api/contratos', methods=['POST'])
@empresa_required
def api_contrato_crear():
    """'Ingresar contrato' — Escenario A (Contratista Minero). El contrato cuelga de la
    empresa activa y hereda su razón social si no se envía otra."""
    f = request.get_json(silent=True) or request.form
    emp = db.empresa_de(session['rut'], _empresa_id()) or {}
    empresa = (f.get('empresa') or '').strip() or emp.get('razon_social') or ''
    numero = (f.get('numero') or '').strip()
    if not empresa or not numero:
        return jsonify({'error': 'Empresa y N° de contrato son obligatorios.'}), 400
    datos = f.get('datos') or {}
    datos_json = json.dumps(datos, ensure_ascii=False) if datos else None
    es_minera = 1 if str(f.get('es_contratista_minera') or '').lower() in ('1', 'si', 'sí', 'true') else 0
    mandante = (f.get('mandante') or '').strip() if es_minera else ''
    cid = db.crear_contrato(session['rut'], empresa, (f.get('faena') or '').strip(),
                            numero, mandante, datos_json, es_contratista_minera=es_minera,
                            empresa_id=_empresa_id())
    return jsonify({'contratos': _consolidar(session['rut']), 'id': cid,
                    'es_contratista_minera': es_minera})


@app.route('/api/contratos/eliminar', methods=['POST'])
@empresa_required
def api_contrato_eliminar():
    f = request.get_json(silent=True) or request.form
    db.eliminar_contrato(session['rut'], int(f.get('id')))
    return jsonify(_consolidar(session['rut']))


# Mandantes mineros parametrizados (selector de derivación / Módulo Puente).
MANDANTES_MINEROS = ['Codelco División RT', 'Minera Spence (BHP)', 'Minera El Abra',
                     'Minera Centinela', 'Otra minería']
FUF_TOTAL = 60          # ítems del FUF DS 44 (base legal Ley 16.744, común a todo empleador)
CARPETA_TOTAL = 29      # ítems de la Carpeta de Arranque (estándar minero)


def _gap_analysis(rut, cid):
    """Compara la base DS 44 (FUF del asesor, ya cumplida) contra el estándar minero
    pendiente (Carpeta de Arranque + RESSO). Devuelve el % de base ya lograda y qué
    falta para alcanzar el estándar minero. El FUF 'suma', no se re-hace."""
    _ce = db.contrato_de(rut, cid) or {}
    fuf = db.estados_fuf(_ce.get('empresa_id') or _empresa_id())
    fuf_ok = sum(1 for r in fuf.values() if r.get('estado') in ('si', 'na'))
    fuf_pct = round(100 * fuf_ok / FUF_TOTAL) if FUF_TOTAL else 0
    car = _carpeta(cid)
    cumple = sum(1 for i in car['items'] if i['estado'] == 'cumple')
    na = sum(1 for i in car['items'] if i['estado'] == 'na')
    pendientes = len(car['items']) - cumple - na
    return {
        'base_ds44': {'pct': fuf_pct, 'cumplidos': fuf_ok, 'total': FUF_TOTAL,
                      'titulo': 'Base legal DS 44 / FUF (Ley 16.744)'},
        'estandar_minero': {
            'carpeta_pct': car['cumplimiento_pct'], 'carpeta_cumple': cumple,
            'carpeta_na': na, 'carpeta_pendientes': pendientes,
            'carpeta_total': len(car['items']),
            'resso_estado': (db.contrato_de(rut, cid) or {}).get('resso_estado') or 'bloqueado'},
        'mensaje': (f'Ya tienes el {fuf_pct}% de la base legal (DS 44) lista. '
                    f'Para el estándar minero faltan {pendientes} ítem(es) de la '
                    f'Carpeta de Arranque y aprobar el RESSO.')
    }


@app.route('/api/contratos/<int:cid>/upgrade', methods=['POST'])
@login_required
def api_contrato_upgrade(cid):
    """Módulo Puente: convierte una empresa general en contratista minera SIN
    re-ingresar datos ni perder avance. Reutiliza el contrato + el FUF del asesor;
    solo fija el mandante y habilita el flujo Carpeta/RESSO. Deriva a la Carpeta."""
    rut = session['rut']
    if not db.contrato_de(rut, cid):
        return jsonify({'error': 'Contrato no encontrado.'}), 404
    f = request.get_json(silent=True) or request.form
    mandante = (f.get('mandante') or '').strip()
    if not mandante:
        return jsonify({'error': 'Selecciona el mandante (minería).'}), 400
    actualizado = db.upgrade_a_contratista_minera(rut, cid, mandante)
    if not actualizado:
        return jsonify({'error': 'No se pudo convertir el contrato.'}), 400
    return jsonify({'contratos': _consolidar(rut), 'id': cid,
                    'gap': _gap_analysis(rut, cid),
                    'es_codelco': resso.es_codelco(mandante)})


@app.route('/api/contratos/<int:cid>/gap', methods=['GET'])
@login_required
def api_contrato_gap(cid):
    rut = session['rut']
    if not db.contrato_de(rut, cid):
        return jsonify({'error': 'Contrato no encontrado.'}), 404
    return jsonify(_gap_analysis(rut, cid))


@app.route('/api/contratos/<int:cid>/matriz', methods=['POST'])
@login_required
def api_subir_matriz(cid):
    rut = session['rut']
    if not db.contrato_de(rut, cid):
        return jsonify({'error': 'Contrato no encontrado.'}), 404
    archivo = request.files.get('archivo')
    if not archivo or not archivo.filename:
        return jsonify({'error': 'No se recibió archivo.'}), 400
    detectados = normativa.parsear_matriz(archivo)
    db.registrar_documento(cid, archivo.filename, request.form.get('flujo', ''), 'matriz')
    aplicados = []
    for d in detectados:
        if d['conforme']:
            db.set_estado_control(rut, cid, d['control_key'], 'aprobado', cid)
            replicar_controles(rut, d['control_key'], cid)
            aplicados.append(normativa.CONTROL_LABEL[d['control_key']])
    return jsonify({'contratos': _consolidar(rut), 'aplicados': aplicados})


# ────────────── Herencia de documentos (Fuente Única de Verdad) ─────────────
def link_document(doc_id, target_contract_id, rut=None):
    """Crea una referencia simbólica (sin copiar archivo) al documento maestro
    `doc_id` dentro del contrato destino. Conserva metadatos de trazabilidad."""
    rut = rut or session['rut']
    master = db.documento_por_id(rut, doc_id)
    if not master:
        return None
    if not db.contrato_de(rut, target_contract_id):
        return None
    if db.existe_referencia(target_contract_id, doc_id):
        return None
    return db.crear_doc_referencia(target_contract_id, master, tipo='matriz')


def sync_to_resso(rut, cid):
    """Recorre los documentos aprobados de la Carpeta de Arranque y crea referencias
    simbólicas en los puntos equivalentes de la Auditoría RESSO (sin duplicar archivos)."""
    docs_item = db.docs_por_item(cid)          # {item_n: [docs...]}
    sincronizados = 0
    for categoria, m in resso.EQUIVALENCIAS.items():
        item_n = m['carpeta']
        base = next((d for d in docs_item.get(item_n, []) if d.get('tipo') == 'evidencia'), None)
        if not base:
            continue
        # el documento del arranque pasa a ser el maestro de su categoría
        db.set_doc_maestro(base['id'], categoria)
        if db.existe_referencia(cid, base['id']):
            continue
        master = db.documento_por_id(rut, base['id'])
        db.registrar_documento(
            cid, master['nombre'], 'RESO ' + m['reso'], 'auditoria_ref',
            item_n=item_n, categoria=categoria, ref_doc_id=base['id'],
            version=master.get('version') or 'v1',
            fecha_aprobacion=master.get('fecha_aprobacion') or date.today().isoformat(),
            firma=master.get('firma'))
        # el punto RESO queda 'cumple' heredado del arranque
        db.set_auditoria_estado(cid, m['reso'], 'cumple',
                                f'Heredado de Carpeta de Arranque (ítem {item_n}).')
        sincronizados += 1
    return sincronizados


@app.route('/api/contratos/<int:cid>/arranque/aprobar', methods=['POST'])
@login_required
def api_arranque_aprobar(cid):
    """Hito ARRANQUE_APROBADO: exige Carpeta al 100% + firmas; activa RESSO y sincroniza."""
    rut = session['rut']
    contrato = db.contrato_de(rut, cid)
    if not contrato:
        return jsonify({'error': 'Contrato no encontrado.'}), 404
    if not resso.es_codelco(contrato.get('mandante')):
        return jsonify({'error': 'El módulo RESSO aplica solo a contratos Codelco.'}), 400
    carp = _carpeta(cid)
    if carp['cumplimiento_pct'] < 100:
        return jsonify({'error': 'La Carpeta de Arranque debe estar al 100% (todos los ítems aprobados) para aprobar el arranque.'}), 400
    f = request.get_json(silent=True) or {}
    docs = db.documentos_de(cid)
    tiene_firmas = f.get('firmas_ok') is True or any(d.get('tipo') == 'firma' for d in docs)
    if not tiene_firmas:
        return jsonify({'error': 'Debe cargar las firmas oficiales (acta de aprobación / Anexo 1) antes de aprobar el arranque.'}), 400
    db.set_arranque_aprobado(cid)
    n = sync_to_resso(rut, cid)
    return jsonify({'ok': True, 'resso_estado': 'en_progreso', 'sincronizados': n,
                    'aviso': 'Se han pre-cargado documentos desde la Carpeta de Arranque aprobada para el RESO'})


def _auditoria(cid):
    """Estado del módulo Auditoría RESSO (puntos + docs heredados por categoría)."""
    estados = db.estados_auditoria(cid)
    por_cat = {}
    for d in db.documentos_de(cid):
        if d.get('tipo') == 'auditoria_ref' and d.get('categoria'):
            por_cat.setdefault(d['categoria'], []).append(d)
    puntos = []
    for p in resso.auditoria_lista():
        e = estados.get(p['punto_key'], {})
        refs = por_cat.get(p['categoria'], [])
        puntos.append({**p,
                       'estado': e.get('estado', 'pendiente'),
                       'observacion': e.get('observacion', '') or '',
                       'docs': [{'nombre': x['nombre'], 'ref_doc_id': x.get('ref_doc_id'),
                                 'version': x.get('version'), 'fecha_aprobacion': x.get('fecha_aprobacion'),
                                 'firma': x.get('firma')} for x in refs]})
    aplicables = [x for x in puntos if x['estado'] != 'na']
    cumple = [x for x in aplicables if x['estado'] == 'cumple']
    pct = round(len(cumple) / len(aplicables) * 100) if aplicables else 0
    return {'puntos': puntos, 'cumplimiento_pct': pct}


@app.route('/api/contratos/<int:cid>/auditoria', methods=['GET'])
@login_required
def api_auditoria(cid):
    c = db.contrato_de(session['rut'], cid)
    if not c:
        return jsonify({'error': 'Contrato no encontrado.'}), 404
    return jsonify({'resso_estado': c.get('resso_estado') or 'bloqueado',
                    'arranque_aprobado': bool(c.get('arranque_aprobado')),
                    **_auditoria(cid)})


# ─────────────── Motor lingüístico: vocabulario técnico + corrección ───────
@app.route('/api/vocabulario', methods=['GET'])
@login_required
def api_vocabulario_listar():
    return jsonify(db.vocabulario_listar(solo_activos=False))


@app.route('/api/vocabulario', methods=['POST'])
@login_required
def api_vocabulario_crear():
    f = request.get_json(silent=True) or {}
    termino = (f.get('termino') or '').strip()
    if not termino:
        return jsonify({'error': 'Indica el término o sigla.'}), 400
    db.vocabulario_crear(termino, (f.get('tipo') or 'termino').strip(),
                         (f.get('significado') or '').strip())
    return jsonify(db.vocabulario_listar(solo_activos=False))


@app.route('/api/vocabulario/<int:vid>/eliminar', methods=['POST'])
@login_required
def api_vocabulario_eliminar(vid):
    db.vocabulario_eliminar(vid)
    return jsonify(db.vocabulario_listar(solo_activos=False))


@app.route('/api/corregir', methods=['POST'])
@login_required
def api_corregir():
    """Corrige ortografía/gramática priorizando el vocabulario técnico. Best-effort:
    nunca lanza 500; ante cualquier fallo devuelve el texto original con un flag."""
    f = request.get_json(silent=True) or {}
    texto = f.get('texto') or ''
    try:
        vocab = db.vocabulario_listar(solo_activos=True)
        return jsonify(correccion.corregir_texto(texto, vocab))
    except Exception:                            # noqa: BLE001 — estabilidad ante todo
        return jsonify({'ok': False, 'original': texto, 'corregido': texto,
                        'cambios': [], 'error': 'No se pudo corregir el texto.'})


@app.route('/api/contratos/<int:cid>/documento', methods=['POST'])
@login_required
def api_registrar_doc(cid):
    rut = session['rut']
    if not db.contrato_de(rut, cid):
        return jsonify({'error': 'Contrato no encontrado.'}), 404
    f = request.get_json(silent=True) or request.form
    nombre = (f.get('nombre') or '').strip()
    if not nombre:
        return jsonify({'error': 'Falta el nombre del documento.'}), 400
    db.registrar_documento(cid, nombre, (f.get('flujo') or '').strip(), 'evidencia')
    return jsonify(_consolidar(rut))


# ─────────────────────── Carpeta de Arranque (RESSO Anexo 2) ───────────────
@app.route('/api/contratos/<int:cid>/carpeta', methods=['GET'])
@login_required
def api_carpeta(cid):
    c = db.contrato_de(session['rut'], cid)
    if not c:
        return jsonify({'error': 'Contrato no encontrado.'}), 404
    if not resso.es_codelco(c['mandante']):
        return jsonify({'error': 'La Carpeta de Arranque RESSO aplica solo a contratos Codelco.'}), 400
    return jsonify(_carpeta(cid))


@app.route('/api/contratos/<int:cid>/carpeta/<int:n>/estado', methods=['POST'])
@login_required
def api_carpeta_estado(cid, n):
    contrato = db.contrato_de(session['rut'], cid)
    if not contrato:
        return jsonify({'error': 'Contrato no encontrado.'}), 404
    f = request.get_json(silent=True) or request.form
    actual = db.estados_carpeta(cid).get(n, {})
    # conservar lo existente cuando el campo no viene en la petición
    estado = f.get('estado')
    if estado is None:
        estado = actual.get('estado', 'pendiente')
    estado = (estado or 'pendiente').strip()
    if estado not in ('pendiente', 'cumple', 'na'):
        estado = 'pendiente'
    obs = f.get('observacion')
    if obs is None:
        obs = actual.get('observacion', '') or ''
    obs = obs.strip()
    fecha_comp = f.get('fecha_compromiso')
    if fecha_comp is None:
        fecha_comp = actual.get('fecha_compromiso')
    fecha_comp = (fecha_comp or '').strip() or None
    # Una brecha (Pendiente) exige observación.
    if estado == 'pendiente' and not obs:
        return jsonify({'error': 'La observación es obligatoria para una brecha (Pendiente).'}), 400
    db.set_item_estado(cid, n, estado, obs, fecha_comp)
    # Carta de No Aplica: se genera/actualiza cuando el ítem queda en N/A; se retira si no.
    if estado == 'na':
        generar_carta_na(session['rut'], cid, contrato, n, obs)
    else:
        db.eliminar_doc_tipo(cid, n, 'carta_na')
    return jsonify(_carpeta(cid))


# ─────────────────────────── FUF DS 44 (persistencia) ──────────────────────
@app.route('/api/fuf', methods=['GET'])
@empresa_required
def api_fuf_get():
    return jsonify(db.estados_fuf(_empresa_id()))


@app.route('/api/fuf', methods=['POST'])
@empresa_required
def api_fuf_guardar():
    """Guarda en bloque los ítems del FUF de la empresa. Exige observación si es 'No Cumple'."""
    f = request.get_json(silent=True) or {}
    items = f.get('items') or []
    rut = session['rut']
    eid = _empresa_id()
    for it in items:
        try:
            n = int(it.get('item_n'))
        except (TypeError, ValueError):
            continue
        estado = (it.get('estado') or 'pendiente').strip()
        if estado not in ('si', 'no', 'na', 'pendiente'):
            estado = 'pendiente'
        obs = (it.get('observacion') or '').strip()
        if estado == 'no' and not obs:
            return jsonify({'error': f'La observación es obligatoria en el ítem {n} (No Cumple).'}), 400
        fecha_comp = (it.get('fecha_compromiso') or '').strip() or None
        db.set_fuf_estado(eid, n, estado, obs, fecha_comp, rut=rut)
    return jsonify(db.estados_fuf(eid))


# ─────────────────────────── Panel de Brechas (unificado) ──────────────────
def _dias_restantes(fecha_iso):
    if not fecha_iso:
        return None
    try:
        d = date.fromisoformat(fecha_iso)
    except (TypeError, ValueError):
        return None
    return (d - date.today()).days


@app.route('/api/brechas', methods=['GET'])
@empresa_required
def api_brechas():
    rut = session['rut']
    eid = _empresa_id()
    brechas = []
    for b in db.brechas_carpeta(rut, eid):
        item = resso.CARPETA_DICT.get(b['item_n'], {})
        brechas.append({
            'fuente': 'carpeta',
            'item_n': b['item_n'],
            'etiqueta': f"Carpeta N°{b['item_n']:02d}",
            'titulo': item.get('titulo', ''),
            'contrato_id': b['contrato_id'],
            'contrato': f"{b['empresa']} · N° {b['numero']}" + (f" · {b['faena']}" if b['faena'] else ''),
            'observacion': b['observacion'] or '',
            'fecha_compromiso': b['fecha_compromiso'] or '',
            'dias_restantes': _dias_restantes(b['fecha_compromiso']),
        })
    for b in db.brechas_fuf(eid):
        brechas.append({
            'fuente': 'fuf',
            'item_n': b['item_n'],
            'etiqueta': f"FUF DS44 N°{b['item_n']}",
            'titulo': '',
            'contrato_id': None,
            'contrato': '—',
            'observacion': b['observacion'] or '',
            'fecha_compromiso': b['fecha_compromiso'] or '',
            'dias_restantes': _dias_restantes(b['fecha_compromiso']),
        })
    return jsonify(brechas)


@app.route('/api/brechas/compromiso', methods=['POST'])
@login_required
def api_brecha_compromiso():
    f = request.get_json(silent=True) or {}
    fuente = f.get('fuente')
    try:
        n = int(f.get('item_n'))
    except (TypeError, ValueError):
        return jsonify({'error': 'item_n inválido.'}), 400
    fecha = (f.get('fecha_compromiso') or '').strip() or None
    if fuente == 'fuf':
        db.set_fuf_compromiso(_empresa_id(), n, fecha)
    elif fuente == 'carpeta':
        cid = f.get('contrato_id')
        if not db.contrato_de(session['rut'], cid):
            return jsonify({'error': 'Contrato no encontrado.'}), 404
        db.set_carpeta_compromiso(cid, n, fecha)
    else:
        return jsonify({'error': 'fuente inválida.'}), 400
    return jsonify({'ok': True})


# ══════════════ Motor de Cumplimiento: pendientes / matriz / reglas ═════════
@app.route('/api/pendientes', methods=['GET'])
@empresa_required
def api_pendientes():
    """Panel de Actividades Pendientes unificado (legal + contractual + operativa),
    con el seguimiento de actualización anual adjunto a cada ítem legal."""
    eid = _empresa_id()
    items = alertas.actividades_pendientes(db, eid)
    seg = db.seguimiento_get(eid)
    for it in items:
        if it.get('tipo') == 'legal' and it.get('categoria') in seg:
            it['seguimiento'] = seg[it['categoria']]
    return jsonify(items)


@app.route('/api/pendientes/<categoria>/seguimiento', methods=['POST'])
@empresa_required
def api_pendiente_seguimiento(categoria):
    """Guarda el comentario y la fecha de compromiso de seguimiento de un documento anual."""
    f = request.get_json(silent=True) or {}
    res = db.seguimiento_set(_empresa_id(), categoria,
                             comentario=(f.get('comentario') or '').strip(),
                             fecha_compromiso=(f.get('fecha_compromiso') or '').strip() or None)
    return jsonify({'ok': True, 'seguimiento': res})


@app.route('/api/pendientes/<categoria>/subir', methods=['POST'])
@empresa_required
def api_pendiente_subir(categoria):
    """Sube la nueva versión de un documento de la Capa Legal, aplica su regla (vencimiento),
    dispara la CASCADA a los contratos mineros y marca el ítem FUF. Doble cobertura."""
    if categoria not in cumplimiento.REGLAS_CUMPLIMIENTO:
        return jsonify({'error': 'Categoría no reconocida.'}), 400
    rut, eid = session['rut'], _empresa_id()
    emp = db.empresa_de(rut, eid) or {}
    fecha_aprob = (request.form.get('fecha_aprobacion') or '').strip()
    if not fecha_aprob:
        return jsonify({'error': 'Indica la fecha de aprobación/emisión del documento.'}), 400
    archivo = request.files.get('archivo')
    contenido = archivo.read() if archivo and archivo.filename else None
    nombre = (archivo.filename if archivo and archivo.filename else
              f"{cumplimiento.REGLAS_CUMPLIMIENTO[categoria]['titulo']} {fecha_aprob[:4]}")
    base_cid = db.contrato_base(eid, rut, emp.get('razon_social'))
    doc_id = db.registrar_documento_legal(
        base_cid, categoria, nombre, fecha_aprob,
        contenido=contenido, mimetype=(archivo.mimetype if archivo else None))
    # Cascada Capa Core → Capa Mandante (referencias a contratos mineros)
    afectados = db.cascada_a_contratos(eid, categoria, doc_id)
    # Nutre el FUF (marca el ítem como cumplido) si la regla lo mapea
    fuf_item = cumplimiento.REGLAS_CUMPLIMIENTO[categoria].get('fuf_item')
    if fuf_item:
        db.set_fuf_estado(eid, fuf_item, 'si', rut=rut)
    mand = ', '.join(sorted({a['mandante'] for a in afectados if a['mandante']})) or 'tus contratos'
    mensaje = (f"Documento actualizado. Cubres tu obligación legal (DS 44) y tu exigencia "
               f"contractual con {mand} simultáneamente"
               + (f"; ítem FUF N°{fuf_item} marcado como cumplido." if fuf_item else "."))
    return jsonify({'ok': True, 'doc_id': doc_id, 'afectados': afectados,
                    'mensaje': mensaje, 'pendientes': alertas.actividades_pendientes(db, eid)})


@app.route('/api/matriz-legal', methods=['GET'])
@empresa_required
def api_matriz_get():
    return jsonify(db.matriz_legal(_empresa_id()))


@app.route('/api/matriz-legal', methods=['POST'])
@empresa_required
def api_matriz_guardar():
    data = request.get_json(silent=True) or {}
    # Capa Operativa: si es fila nueva sin ID, genera uno (OP-N) automáticamente.
    if not (data.get('id_requisito') or '').strip():
        existentes = db.matriz_legal(_empresa_id())
        n = 1 + sum(1 for r in existentes if str(r.get('id_requisito') or '').startswith('OP-'))
        data['id_requisito'] = f'OP-{n:03d}'
        data.setdefault('capa', 'operativa')
    return jsonify(db.requisito_guardar(_empresa_id(), data))


@app.route('/api/matriz-legal/<int:rid>/eliminar', methods=['POST'])
@empresa_required
def api_matriz_eliminar(rid):
    res = db.requisito_eliminar(_empresa_id(), rid)
    if res.get('error'):
        return jsonify(res), 400
    return jsonify(db.matriz_legal(_empresa_id()))


@app.route('/api/matriz-legal/importar', methods=['POST'])
@empresa_required
def api_matriz_importar():
    """Importa la Matriz Legal desde un Excel normalizado (columnas ID_Requisito, Origen,
    Cuerpo_Normativo, Requisito_Legal, Riesgo_Asociado, Control_Operativo, Frecuencia,
    Estado_Avance). Reusa openpyxl."""
    archivo = request.files.get('archivo')
    if not archivo or not archivo.filename:
        return jsonify({'error': 'No se recibió archivo.'}), 400
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(archivo.read()), data_only=True)
        ws = wb.active
        filas = list(ws.iter_rows(values_only=True))
    except Exception as e:      # noqa: BLE001
        return jsonify({'error': f'No se pudo leer el Excel: {type(e).__name__}.'}), 400
    if not filas:
        return jsonify({'error': 'El archivo está vacío.'}), 400
    encabezados = [str(h or '').strip().lower().replace(' ', '_') for h in filas[0]]
    campo = {'id_requisito': 'id_requisito', 'origen': 'origen', 'cuerpo_normativo': 'cuerpo_normativo',
             'requisito_legal': 'requisito_legal', 'riesgo_asociado': 'riesgo_asociado',
             'control_operativo': 'control_operativo', 'frecuencia': 'frecuencia',
             'estado_avance': 'estado_avance', 'responsable': 'responsable', 'capa': 'capa',
             'categoria': 'categoria'}
    n = 0
    for fila in filas[1:]:
        data = {}
        for i, h in enumerate(encabezados):
            if h in campo and i < len(fila) and fila[i] is not None:
                data[campo[h]] = str(fila[i]).strip()
        if not data:
            continue
        db.requisito_guardar(_empresa_id(), data)
        n += 1
    return jsonify({'ok': True, 'importados': n, 'matriz': db.matriz_legal(_empresa_id())})


@app.route('/api/reglas', methods=['GET'])
@login_required
def api_reglas_get():
    return jsonify(db.reglas_listar())


@app.route('/api/reglas/<int:rid>', methods=['POST'])
@login_required
def api_regla_editar(rid):
    f = request.get_json(silent=True) or {}
    r = db.regla_actualizar(rid, periodicidad_meses=f.get('periodicidad_meses'),
                            es_critico=f.get('es_critico'))
    if not r:
        return jsonify({'error': 'Regla no encontrada.'}), 404
    return jsonify(r)


# ══════════════ Ronda 15 — Motor Documental: Trabajadores / Tareas / EPP / PTS / IRL ═════════
def _logo_empresa(rut, empresa_id):
    """Logo de la empresa: primero el logo corporativo (empresa.logo_doc_id); si no, el de
    cualquiera de sus contratos. Devuelve un data URI, o None."""
    emp = db.empresa_de(rut, empresa_id) or {}
    if emp.get('logo_doc_id'):
        blob = db.documento_contenido(rut, emp['logo_doc_id'])
        if blob:
            import base64
            contenido, mimetype, _ = blob
            return f"data:{mimetype or 'image/png'};base64,{base64.b64encode(contenido).decode()}"
    for c in db.listar_contratos(rut, empresa_id):
        uri = _logo_data_uri(rut, c['id'])
        if uri:
            return uri
    return None


@app.route('/api/empresa/logo', methods=['POST'])
@empresa_required
def api_empresa_logo():
    """Carga/reemplaza el logo corporativo de la empresa activa (blob en la BD)."""
    rut, eid = session['rut'], _empresa_id()
    emp = db.empresa_de(rut, eid)
    if not emp:
        return jsonify({'error': 'Empresa no encontrada.'}), 404
    archivo = request.files.get('logo')
    if not archivo or not archivo.filename:
        return jsonify({'error': 'No se recibió imagen.'}), 400
    ext = os.path.splitext(archivo.filename)[1].lower() or '.png'
    base_cid = db.contrato_base(eid, rut, emp.get('razon_social'))
    db.eliminar_doc_tipo(base_cid, None, 'logo_empresa')      # reemplazar anterior
    doc_id = db.registrar_documento(base_cid, 'logo_empresa' + ext, '', 'logo_empresa',
                                    contenido=archivo.read(), mimetype=archivo.mimetype or 'image/png')
    db.empresa_set_logo(eid, doc_id)
    return jsonify({'ok': True, 'logo_doc_id': doc_id, 'url': f'/api/doc/{doc_id}'})


@app.route('/api/empresa/logo', methods=['GET'])
@empresa_required
def api_empresa_logo_get():
    emp = db.empresa_de(session['rut'], _empresa_id()) or {}
    return jsonify({'logo_doc_id': emp.get('logo_doc_id'),
                    'url': f"/api/doc/{emp['logo_doc_id']}" if emp.get('logo_doc_id') else None,
                    'razon_social': emp.get('razon_social')})


# ── Trabajadores ──
@app.route('/api/trabajadores', methods=['GET'])
@empresa_required
def api_trabajadores():
    out = []
    for t in db.trabajadores_de(_empresa_id()):
        t['tareas'] = [x['id'] for x in db.tareas_de_trabajador(t['id'])]
        t['irls'] = db.irls_de_trabajador(t['id'])
        out.append(t)
    return jsonify(out)


@app.route('/api/trabajadores', methods=['POST'])
@empresa_required
def api_trabajador_crear():
    f = request.get_json(silent=True) or request.form
    rut = (f.get('rut') or '').strip()
    nombre = (f.get('nombre') or '').strip()
    if not (rut and nombre):
        return jsonify({'error': 'Indica RUT y nombre del trabajador.'}), 400
    if not rut_valido(rut):
        return jsonify({'error': 'El RUT del trabajador no es válido.'}), 400
    if len(db.trabajadores_de(_empresa_id())) >= MAX_TRABAJADORES_BASICO:
        return jsonify({'error': 'limite_trabajadores',
                        'mensaje': f'Alcanzaste el tope del Plan Básico ({MAX_TRABAJADORES_BASICO} '
                                   'trabajadores por empresa). Migra a un plan superior (25-50 '
                                   'trabajadores) para agregar más.'}), 403
    db.trabajador_crear(_empresa_id(), normalizar_rut(rut), nombre,
                        cargo=(f.get('cargo') or '').strip() or None,
                        rol=(f.get('rol') or '').strip() or None)
    return jsonify(db.trabajadores_de(_empresa_id()))


@app.route('/api/trabajadores/<int:tid>/eliminar', methods=['POST'])
@empresa_required
def api_trabajador_eliminar(tid):
    db.trabajador_eliminar(_empresa_id(), tid)
    return jsonify(db.trabajadores_de(_empresa_id()))


@app.route('/api/trabajadores/<int:tid>/tareas', methods=['POST'])
@empresa_required
def api_trabajador_tareas(tid):
    if not db.trabajador_de(_empresa_id(), tid):
        return jsonify({'error': 'Trabajador no encontrado.'}), 404
    f = request.get_json(silent=True) or {}
    db.trabajador_set_tareas(tid, f.get('tarea_ids') or [])
    return jsonify({'ok': True, 'tareas': [x['id'] for x in db.tareas_de_trabajador(tid)]})


# ── Tareas / EPP / PTS de la Matriz IPER ──
@app.route('/api/iper/tareas', methods=['GET'])
@empresa_required
def api_tareas_get():
    tareas = db.tareas_de_empresa(_empresa_id())
    for t in tareas:
        t['riesgos'] = db.riesgos_de_tarea(t['id'])
        t['epp'] = db.epp_de_tarea(t['id'])
        t['pts'] = db.pts_de_tarea(t['id'])
    return jsonify(tareas)


@app.route('/api/iper/tareas', methods=['POST'])
@empresa_required
def api_tarea_crear():
    f = request.get_json(silent=True) or request.form
    nombre = (f.get('nombre') or '').strip()
    if not nombre:
        return jsonify({'error': 'Indica el nombre de la Tarea.'}), 400
    eid = _empresa_id()
    m = db.matriz_riesgo_vigente(eid) or {'id': db.crear_matriz_riesgo(eid, session.get('nombre'))}
    tid = db.tarea_crear(m['id'], nombre, proceso=(f.get('proceso') or '').strip() or None,
                         rutinaria=(f.get('rutinaria') or '').strip() or None,
                         responsable=(f.get('responsable') or '').strip() or None)
    # riesgos opcionales en el alta → se amarran a la tarea recién creada
    from models import RiesgoItem
    for r in (f.get('riesgos') or []):
        rid = db.riesgo_agregar(m['id'], r.get('peligro'), r.get('riesgo'), r.get('medida_control'),
                                nivel_riesgo=r.get('nivel_riesgo'), es_critico=r.get('es_critico', 0))
        it = RiesgoItem.query.get(rid)
        if it:
            it.tarea_id = tid
    sqla.session.commit()
    return jsonify({'ok': True, 'tarea_id': tid})


@app.route('/api/iper/tareas/<int:tid>/riesgo', methods=['POST'])
@empresa_required
def api_tarea_riesgo(tid):
    f = request.get_json(silent=True) or {}
    eid = _empresa_id()
    m = db.matriz_riesgo_vigente(eid)
    if not m:
        return jsonify({'error': 'No hay matriz vigente.'}), 400
    rid = db.riesgo_agregar(m['id'], f.get('peligro'), f.get('riesgo'), f.get('medida_control'),
                            nivel_riesgo=f.get('nivel_riesgo'), es_critico=f.get('es_critico', 0))
    # amarrar el riesgo a la tarea
    from models import RiesgoItem
    it = RiesgoItem.query.get(rid)
    if it:
        it.tarea_id = tid
        sqla.session.commit()
    return jsonify({'ok': True, 'riesgo_id': rid})


@app.route('/api/epp', methods=['GET'])
@empresa_required
def api_epp_get():
    return jsonify(db.epp_listar(_empresa_id()))


@app.route('/api/epp', methods=['POST'])
@empresa_required
def api_epp_crear():
    f = request.get_json(silent=True) or request.form
    nombre = (f.get('nombre') or '').strip()
    if not nombre:
        return jsonify({'error': 'Indica el nombre del EPP.'}), 400
    db.epp_crear(_empresa_id(), nombre, codigo=(f.get('codigo') or '').strip() or None,
                 norma=(f.get('norma') or '').strip() or None)
    return jsonify(db.epp_listar(_empresa_id()))


@app.route('/api/pts', methods=['GET'])
@empresa_required
def api_pts_get():
    return jsonify(db.pts_listar(_empresa_id()))


@app.route('/api/pts', methods=['POST'])
@empresa_required
def api_pts_crear():
    f = request.get_json(silent=True) or request.form
    nombre = (f.get('nombre') or '').strip()
    if not nombre:
        return jsonify({'error': 'Indica el nombre del PTS.'}), 400
    db.pts_crear(_empresa_id(), nombre, codigo=(f.get('codigo') or '').strip() or None,
                 version=(f.get('version') or '').strip() or None)
    return jsonify(db.pts_listar(_empresa_id()))


@app.route('/api/iper/tareas/<int:tid>/epp', methods=['POST'])
@empresa_required
def api_tarea_link_epp(tid):
    f = request.get_json(silent=True) or {}
    db.tarea_link_epp(tid, int(f.get('epp_id')))
    return jsonify({'ok': True, 'epp': db.epp_de_tarea(tid)})


@app.route('/api/iper/tareas/<int:tid>/pts', methods=['POST'])
@empresa_required
def api_tarea_link_pts(tid):
    f = request.get_json(silent=True) or {}
    db.tarea_link_pts(tid, int(f.get('pts_id')))
    return jsonify({'ok': True, 'pts': db.pts_de_tarea(tid)})


# ── IRL: generar / actualizar / listar ──
@app.route('/api/irl/generar', methods=['POST'])
@empresa_required
def api_irl_generar():
    f = request.get_json(silent=True) or request.form
    rut, eid = session['rut'], _empresa_id()
    trab = db.trabajador_de(eid, int(f.get('trabajador_id')))
    if not trab:
        return jsonify({'error': 'Trabajador no encontrado.'}), 404
    empresa = db.empresa_de(rut, eid)
    logo = _logo_empresa(rut, eid)
    res = docgen.generar_irl(db, empresa, trab, session.get('nombre') or rut, logo_data_uri=logo)
    return jsonify({'ok': True, 'doc_id': res['doc_id'], 'audit_id': res['audit_id'],
                    'matriz_version': res['matriz_version'],
                    'preview_url': f"/api/doc/{res['doc_id']}"})


@app.route('/api/trabajadores/<int:tid>/irl', methods=['GET'])
@empresa_required
def api_trabajador_irls(tid):
    return jsonify(db.irls_de_trabajador(tid))


# ══════════════ Ronda 17 — Consola MIPER (VEP DS 44) + biblioteca + cascada IRL ══════════
@app.route('/api/miper', methods=['GET'])
@empresa_required
def api_miper():
    """Matriz de riesgos vigente: tareas → riesgos (con VEP/magnitud/método) + contexto minero."""
    eid = _empresa_id()
    m = db.matriz_riesgo_vigente(eid)
    if not m:
        m = {'id': db.crear_matriz_riesgo(eid, session.get('nombre'))}
        m = db.matriz_riesgo_vigente(eid)
    tareas = db.tareas_de_matriz(m['id'])
    for t in tareas:
        t['riesgos'] = db.riesgos_de_tarea(t['id'])
        t['trabajadores'] = [x['id'] for x in db.trabajadores_de_tarea(t['id'])]
    return jsonify({'matriz': m, 'tareas': tareas,
                    'mineros': db.contratos_mineros_de(eid), 'ecf_puntos': iper.ECF_PUNTOS,
                    'escala': {'probabilidad': iper.PROBABILIDAD, 'consecuencia': iper.CONSECUENCIA}})


@app.route('/api/miper/tareas', methods=['POST'])
@empresa_required
def api_miper_tarea():
    """Crea una tarea manual con sus riesgos. Si guardar_biblioteca=1, la persiste para reuso."""
    f = request.get_json(silent=True) or {}
    nombre = (f.get('nombre') or '').strip()
    if not nombre:
        return jsonify({'error': 'Indica el nombre de la tarea.'}), 400
    eid = _empresa_id()
    m = db.matriz_riesgo_vigente(eid) or {'id': db.crear_matriz_riesgo(eid, session.get('nombre'))}
    mid = m['id'] if isinstance(m, dict) else m
    tid = db.tarea_crear(mid, nombre, proceso=(f.get('proceso') or '').strip() or None,
                         responsable=(f.get('responsable') or '').strip() or None,
                         rutinaria=(f.get('rutinaria') or '').strip() or None)
    from models import RiesgoItem
    for r in (f.get('riesgos') or []):
        rid = db.riesgo_agregar(mid, r.get('peligro'), r.get('riesgo'), r.get('medida_control'),
                                probabilidad=r.get('probabilidad'), consecuencia=r.get('consecuencia'),
                                metodo_correcto=r.get('metodo_correcto'),
                                contrato_id=r.get('contrato_id') or None,
                                ecf_punto=r.get('ecf_punto'), mfl=r.get('mfl'), bowtie=r.get('bowtie'))
        it = RiesgoItem.query.get(rid)
        if it:
            it.tarea_id = tid
    sqla.session.commit()
    # auto-aprendizaje: guardar en biblioteca personalizada
    if f.get('guardar_biblioteca'):
        for r in (f.get('riesgos') or [{}]):
            db.biblioteca_crear(eid, nombre, peligro=r.get('peligro'), riesgo=r.get('riesgo'),
                                medida_control=r.get('medida_control'),
                                metodo_correcto=r.get('metodo_correcto'),
                                probabilidad=r.get('probabilidad'), consecuencia=r.get('consecuencia'))
    return jsonify({'ok': True, 'tarea_id': tid})


@app.route('/api/mutuales', methods=['GET'])
@login_required
def api_mutuales():
    return jsonify(iper.MUTUALES)


@app.route('/api/miper/sugerencia', methods=['GET'])
@empresa_required
def api_miper_sugerencia():
    """Sugiere el requisito legal aplicable a una actividad/peligro/riesgo (MIPER → Legal)."""
    s = iper.sugerir_requisito(request.args.get('texto', ''))
    return jsonify({'sugerencia': s})


@app.route('/api/miper/riesgo/<int:rid>/vincular-legal', methods=['POST'])
@empresa_required
def api_miper_vincular_legal(rid):
    """Vincula el riesgo con el requisito legal sugerido (creándolo en la Matriz Legal si falta)."""
    f = request.get_json(silent=True) or {}
    eid = _empresa_id()
    s = f.get('sugerencia') or iper.sugerir_requisito(f.get('texto', ''))
    if not s:
        return jsonify({'error': 'Sin requisito sugerido.'}), 400
    req_row = db.asegurar_requisito_sugerido(eid, s)
    db.vincular_riesgo_requisito(rid, req_row)
    return jsonify({'ok': True, 'requisito_row_id': req_row, 'cuerpo_legal': s.get('cuerpo_legal')})


@app.route('/api/matriz-legal/<int:rid>/riesgos', methods=['GET'])
@empresa_required
def api_matriz_riesgos(rid):
    """Legal → MIPER: actividades/riesgos vinculados a un requisito legal."""
    return jsonify(db.riesgos_de_requisito(_empresa_id(), rid))


@app.route('/api/miper/desde-requisito', methods=['POST'])
@empresa_required
def api_miper_desde_requisito():
    """Legal → MIPER: crea una tarea/riesgo en la MIPER a partir de un requisito legal, ya vinculado."""
    f = request.get_json(silent=True) or {}
    from models import RequisitoLegal, RiesgoItem
    req = RequisitoLegal.query.filter_by(id=f.get('requisito_id'), empresa_id=_empresa_id()).first()
    if not req:
        return jsonify({'error': 'Requisito no encontrado.'}), 404
    eid = _empresa_id()
    m = db.matriz_riesgo_vigente(eid) or {'id': db.crear_matriz_riesgo(eid, session.get('nombre'))}
    mid = m['id'] if isinstance(m, dict) else m
    nombre = (req.requisito_legal or req.cuerpo_normativo or 'Actividad legal')[:60]
    tid = db.tarea_crear(mid, nombre, proceso='Legal')
    rid = db.riesgo_agregar(mid, req.riesgo_asociado or '', req.requisito_legal or '',
                            req.control_operativo or '', probabilidad=2, consecuencia=2)
    it = RiesgoItem.query.get(rid)
    if it:
        it.tarea_id = tid
        it.requisito_legal_id = req.id
        sqla.session.commit()
    return jsonify({'ok': True, 'tarea_id': tid, 'riesgo_id': rid})


@app.route('/api/miper/riesgo', methods=['POST'])
@empresa_required
def api_miper_riesgo_add():
    """Agrega un riesgo a una tarea existente."""
    f = request.get_json(silent=True) or {}
    tid = f.get('tarea_id')
    eid = _empresa_id()
    m = db.matriz_riesgo_vigente(eid)
    if not (m and tid):
        return jsonify({'error': 'Falta matriz o tarea.'}), 400
    rid = db.riesgo_agregar(m['id'], f.get('peligro'), f.get('riesgo'), f.get('medida_control'),
                            probabilidad=f.get('probabilidad'), consecuencia=f.get('consecuencia'),
                            metodo_correcto=f.get('metodo_correcto'),
                            contrato_id=f.get('contrato_id') or None)
    from models import RiesgoItem
    it = RiesgoItem.query.get(rid)
    if it:
        it.tarea_id = int(tid)
        sqla.session.commit()
    return jsonify({'ok': True, 'riesgo_id': rid})


@app.route('/api/miper/riesgo/<int:rid>/campo', methods=['POST'])
@empresa_required
def api_miper_riesgo_campo(rid):
    """Edita en línea un campo del riesgo (recalcula VEP). Si cambia una medida/método/peligro/riesgo,
    dispara la cascada al IRL (Art. 15 DS 44): regenera los IRL de los trabajadores afectados con
    aviso de que requieren nueva firma."""
    f = request.get_json(silent=True) or {}
    campo, valor = f.get('campo'), f.get('valor')
    it, afecta_irl = db.riesgo_editar(rid, campo, valor)
    if it is None:
        return jsonify({'error': 'Campo no editable o riesgo no encontrado.'}), 400
    resp = {'ok': True, 'riesgo': it, 'irls_actualizados': []}
    if afecta_irl and it.get('tarea_id'):
        rut, eid = session['rut'], _empresa_id()
        empresa = db.empresa_de(rut, eid)
        logo = _logo_empresa(rut, eid)
        afectados = docgen.regenerar_irls_de_tarea(
            db, empresa, it['tarea_id'], session.get('nombre') or rut, logo_data_uri=logo,
            motivo=f'Cambio en “{campo}” de la Matriz IPER')
        resp['irls_actualizados'] = afectados
    return jsonify(resp)


@app.route('/api/miper/riesgo/<int:rid>/eliminar', methods=['POST'])
@empresa_required
def api_miper_riesgo_del(rid):
    from models import RiesgoItem
    it = RiesgoItem.query.get(rid)
    m = db.matriz_riesgo_vigente(_empresa_id())
    if it and m and it.matriz_id == m['id']:
        sqla.session.delete(it)
        sqla.session.commit()
    return jsonify({'ok': True})


@app.route('/api/miper/versionar', methods=['POST'])
@empresa_required
def api_miper_versionar():
    f = request.get_json(silent=True) or {}
    motivo = (f.get('motivo') or 'Revisión periódica').strip()
    nueva = db.bloquear_y_versionar(_empresa_id(), motivo, session.get('nombre'))
    return jsonify({'ok': True, 'matriz': nueva})


# ══════════════ Fase 2 — Gestión de faena (precarga por contrato) ══════════════
@app.route('/api/faena/<int:cid>', methods=['GET'])
@empresa_required
def api_faena_get(cid):
    """Capa legal (mandante) + riesgos de la faena de un contrato + catálogo de actividades."""
    rut, eid = session['rut'], _empresa_id()
    c = db.contrato_de(rut, cid)
    if not c:
        return jsonify({'error': 'Contrato no encontrado.'}), 404
    actividades = [t['tarea'] for t in iper.CATALOGO_TAREAS_BASE] + \
                  [b['nombre'] for b in db.biblioteca_listar(eid)]
    return jsonify({'contrato': c,
                    'legal': db.matriz_legal_contrato(eid, cid),
                    'riesgos': db.riesgos_de_contrato(eid, cid),
                    'actividades': sorted(set(actividades))})


@app.route('/api/faena/<int:cid>/precargar', methods=['POST'])
@empresa_required
def api_faena_precargar(cid):
    res = db.precargar_faena(session['rut'], cid)
    if res.get('error'):
        return jsonify(res), 400
    return jsonify({'ok': True, **res,
                    'legal': db.matriz_legal_contrato(_empresa_id(), cid),
                    'riesgos': db.riesgos_de_contrato(_empresa_id(), cid)})


@app.route('/api/faena/<int:cid>/legal', methods=['POST'])
@empresa_required
def api_faena_legal(cid):
    """Agrega un requisito legal de mandante ligado a la faena (CRUD capa mandante)."""
    eid = _empresa_id()
    if not db.contrato_de(session['rut'], cid):
        return jsonify({'error': 'Contrato no encontrado.'}), 404
    f = request.get_json(silent=True) or {}
    existentes = db.matriz_legal_contrato(eid, cid)
    n = 1 + len(existentes)
    data = {'id_requisito': f"F{cid}-OP-{n:02d}", 'capa': 'mandante', 'contrato_id': cid,
            'origen': (f.get('origen') or '').strip(), 'cuerpo_normativo': (f.get('cuerpo_normativo') or '').strip(),
            'requisito_legal': (f.get('requisito_legal') or '').strip(),
            'responsable': (f.get('responsable') or '').strip()}
    db.requisito_guardar(eid, data)
    return jsonify(db.matriz_legal_contrato(eid, cid))


@app.route('/api/faena/<int:cid>/pauta', methods=['GET'])
@empresa_required
def api_faena_pauta(cid):
    """Genera la Pauta de Arranque RESSO (HTML imprimible → PDF): ítems de la Carpeta de Arranque
    + estándares del mandante de la faena, con el logo de la empresa. Reusa el patrón de carta_na."""
    rut, eid = session['rut'], _empresa_id()
    c = db.contrato_de(rut, cid)
    if not c:
        return ('Contrato no encontrado', 404)
    logo = _logo_empresa(rut, eid)
    emp = db.empresa_de(rut, eid) or {}
    carpeta = resso.carpeta_lista() if hasattr(resso, 'carpeta_lista') else []
    if not carpeta:
        carpeta = [{'n': n, 'titulo': t, 'evidencia': e, 'a_quien': q}
                   for (n, t, e, q) in resso.CARPETA_ARRANQUE]
    estandares = db.matriz_legal_contrato(eid, cid)
    hoy = date.today().strftime('%d-%m-%Y')
    def esc(s):
        return (str(s or '')).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    logo_html = f'<img src="{logo}" style="max-height:74px;max-width:220px">' if logo else \
        f'<div style="font-weight:800;color:#006a9b;font-size:20px">{esc(emp.get("razon_social") or "Empresa")}</div>'
    filas_carp = ''.join(
        f'<tr><td style="text-align:center">{it["n"]}</td><td><b>{esc(it["titulo"])}</b><br>'
        f'<span style="color:#666;font-size:11px">{esc(it.get("evidencia"))}</span></td>'
        f'<td>{esc(it.get("a_quien"))}</td><td style="text-align:center">☐</td></tr>' for it in carpeta)
    filas_est = ''.join(
        f'<tr><td>{esc(e.get("id_requisito"))}</td><td>{esc(e.get("requisito_legal"))}</td>'
        f'<td>{esc(e.get("estado_avance"))}</td></tr>' for e in estandares) or \
        '<tr><td colspan="3" style="color:#999">Sin estándares del mandante precargados. Usa “Precargar”.</td></tr>'
    html = f"""<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<title>Pauta de Arranque · {esc(c.get('numero',''))}</title><style>
 body{{font-family:Arial,Helvetica,sans-serif;color:#1a2b3c;max-width:900px;margin:24px auto;padding:0 24px;line-height:1.45;font-size:12px}}
 .head{{display:flex;justify-content:space-between;align-items:center;border-bottom:3px solid #006a9b;padding-bottom:12px;margin-bottom:14px}}
 h1{{font-size:17px;margin:0;color:#006a9b}} .sub{{font-size:11px;color:#666}}
 table{{width:100%;border-collapse:collapse;margin:10px 0}} th,td{{border:1px solid #d1d5db;padding:6px 8px;text-align:left;vertical-align:top}}
 th{{background:#eef2f5;font-size:10px;text-transform:uppercase;color:#374151}} h2{{font-size:13px;color:#006a9b;margin:18px 0 4px}}
 .pie{{margin-top:22px;font-size:10px;color:#999}} @media print{{.noprint{{display:none}}}}
</style></head><body>
 <div class="head">{logo_html}<div style="text-align:right"><h1>PAUTA DE ARRANQUE — CARPETA (RESSO)</h1>
  <div class="sub">Estándar minero · {esc(c.get('mandante',''))}{(' — ' + esc(c.get('faena'))) if c.get('faena') else ''}</div></div></div>
 <table><tr><td class="k" style="width:180px;background:#f4f7f6"><b>Empresa</b></td><td>{esc(emp.get('razon_social'))}</td>
  <td class="k" style="width:140px;background:#f4f7f6"><b>N° Contrato</b></td><td>{esc(c.get('numero',''))}</td></tr>
  <tr><td style="background:#f4f7f6"><b>Mandante</b></td><td>{esc(c.get('mandante',''))}</td>
  <td style="background:#f4f7f6"><b>Fecha</b></td><td>{esc(hoy)}</td></tr></table>
 <h2>1. Carpeta de Arranque (29 ítems)</h2>
 <table><thead><tr><th style="width:36px">N°</th><th>Documento / evidencia</th><th style="width:150px">Responsable</th><th style="width:44px">✔</th></tr></thead><tbody>{filas_carp}</tbody></table>
 <h2>2. Estándares del mandante (RC · ECF · SST)</h2>
 <table><thead><tr><th style="width:90px">Código</th><th>Estándar</th><th style="width:110px">Estado</th></tr></thead><tbody>{filas_est}</tbody></table>
 <div class="pie">Smart HSE Chile · Pauta generada automáticamente desde la Carpeta de Arranque y los estándares del mandante.
  <button class="noprint" onclick="window.print()">Imprimir / Guardar PDF</button></div>
</body></html>"""
    return send_file(BytesIO(html.encode('utf-8')), mimetype='text/html', as_attachment=False,
                     download_name=f'Pauta_Arranque_{re.sub(r"[^\\w-]+","_",c.get("numero","") or "")}.html')


@app.route('/api/faena/<int:cid>/inyectar', methods=['POST'])
@empresa_required
def api_faena_inyectar(cid):
    """Inyecta actividades/procesos (del catálogo base o biblioteca) a la MIPER de la faena."""
    f = request.get_json(silent=True) or {}
    res = db.inyectar_actividades_faena(session['rut'], cid, f.get('actividades') or [])
    if res.get('error'):
        return jsonify(res), 400
    return jsonify({'ok': True, **res, 'riesgos': db.riesgos_de_contrato(_empresa_id(), cid)})


@app.route('/api/biblioteca', methods=['GET'])
@empresa_required
def api_biblioteca_get():
    return jsonify(db.biblioteca_listar(_empresa_id()))


@app.route('/api/biblioteca', methods=['POST'])
@empresa_required
def api_biblioteca_add():
    f = request.get_json(silent=True) or {}
    nombre = (f.get('nombre') or '').strip()
    if not nombre:
        return jsonify({'error': 'Indica el nombre.'}), 400
    db.biblioteca_crear(_empresa_id(), nombre, peligro=f.get('peligro'), riesgo=f.get('riesgo'),
                        medida_control=f.get('medida_control'), metodo_correcto=f.get('metodo_correcto'),
                        probabilidad=f.get('probabilidad'), consecuencia=f.get('consecuencia'))
    return jsonify(db.biblioteca_listar(_empresa_id()))


@app.route('/api/contratos/<int:cid>/carpeta/<int:n>/documento', methods=['POST'])
@login_required
def api_carpeta_doc(cid, n):
    c = db.contrato_de(session['rut'], cid)
    if not c:
        return jsonify({'error': 'Contrato no encontrado.'}), 404
    archivos = request.files.getlist('archivo') or []
    if not archivos:
        return jsonify({'error': 'No se recibió archivo.'}), 400
    for archivo in archivos:
        if not archivo or not archivo.filename:
            continue
        db.registrar_documento(cid, archivo.filename, '', 'evidencia', item_n=n,
                               contenido=archivo.read(),
                               mimetype=archivo.mimetype or 'application/octet-stream')
    return jsonify(_carpeta(cid))


@app.route('/api/contratos/<int:cid>/carpeta/bulk', methods=['POST'])
@login_required
def api_carpeta_bulk(cid):
    """Carga masiva: varios archivos / carpeta completa → se clasifican en los 29 ítems."""
    c = db.contrato_de(session['rut'], cid)
    if not c:
        return jsonify({'error': 'Contrato no encontrado.'}), 404
    archivos = request.files.getlist('archivos') or []
    if not archivos:
        return jsonify({'error': 'No se recibieron archivos.'}), 400
    reporte = []
    for archivo in archivos:
        if not archivo or not archivo.filename:
            continue
        # webkitdirectory envía rutas relativas; respetar numeración de carpeta si existe
        nombre = archivo.filename.replace('\\', '/').split('/')[-1]
        n, fuente = ia.clasificar_path(archivo.filename)
        db.registrar_documento(cid, nombre, '', 'evidencia', item_n=n,
                               contenido=archivo.read(),
                               mimetype=archivo.mimetype or 'application/octet-stream')
        reporte.append({'archivo': nombre, 'item': n,
                        'titulo': resso.CARPETA_DICT.get(n, {}).get('titulo', ''),
                        'fuente': fuente})
    return jsonify({'carpeta': _carpeta(cid), 'reporte': reporte})


@app.route('/api/doc/<int:doc_id>', methods=['GET'])
@login_required
def api_doc(doc_id):
    """Sirve un documento (evidencia, carta o logo) desde la BD, resolviendo referencias."""
    blob = db.documento_contenido(session['rut'], doc_id)
    if not blob:
        return ('Documento no encontrado', 404)
    contenido, mimetype, nombre = blob
    inline = (mimetype or '').startswith(('image/', 'text/html')) or \
        (nombre or '').lower().endswith(('.html', '.htm'))
    return send_file(BytesIO(contenido), mimetype=mimetype or 'application/octet-stream',
                     as_attachment=not inline, download_name=nombre or f'doc_{doc_id}')


@app.route('/api/contratos/<int:cid>/logo', methods=['POST'])
@login_required
def api_subir_logo(cid):
    c = db.contrato_de(session['rut'], cid)
    if not c:
        return jsonify({'error': 'Contrato no encontrado.'}), 404
    archivo = request.files.get('logo')
    if not archivo or not archivo.filename:
        return jsonify({'error': 'No se recibió imagen.'}), 400
    ext = os.path.splitext(archivo.filename)[1].lower() or '.png'
    # reemplazar logo anterior
    for d in db.documentos_de(cid):
        if d.get('tipo') == 'logo':
            db.eliminar_doc_tipo(cid, None, 'logo')
            break
    doc_id = db.registrar_documento(cid, 'empresa_logo' + ext, '', 'logo',
                                    contenido=archivo.read(),
                                    mimetype=archivo.mimetype or 'image/png')
    datos = _datos(c)
    datos['logo_doc_id'] = doc_id                 # referencia al blob en la BD
    datos.pop('logo', None)                       # limpiar esquema antiguo (filesystem)
    db.actualizar_datos(cid, json.dumps(datos, ensure_ascii=False))
    return jsonify(_consolidar(session['rut']))


@app.route('/api/contratos/<int:cid>/logo', methods=['GET'])
@login_required
def api_ver_logo(cid):
    c = db.contrato_de(session['rut'], cid)
    if not c:
        return ('', 404)
    doc = _logo_doc(cid)
    if not doc:
        return ('', 404)
    blob = db.documento_contenido(session['rut'], doc['id'])
    if not blob:
        return ('', 404)
    contenido, mimetype, _ = blob
    return send_file(BytesIO(contenido), mimetype=mimetype or 'image/png')


if __name__ == '__main__':
    app.run(debug=True, port=5000)
